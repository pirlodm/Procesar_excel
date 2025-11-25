import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
import re

# --- CONFIGURACIÓN ---
BASE_DIR = Path(__file__).parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"

# El archivo de salida será uno solo con todo unido
ARCHIVO_SALIDA = "datos_transformados_unificados.xlsx"
RUTA_SALIDA = OUTPUT_DIR / ARCHIVO_SALIDA

# Nombres de hojas esperados
HOJA_DATOS = "RECLAM. FB"
HOJA_METADATOS = "RESULTADOS FB"

def encontrar_producto_en_resultados(wb, hoja_nombre):
    """Busca el producto en la hoja 'RESULTADOS FB', celda C2."""
    producto_detectado = "DESCONOCIDO"
    if hoja_nombre not in wb.sheetnames:
        return producto_detectado
    
    celda_c2 = wb[hoja_nombre]["C2"].value
    if celda_c2 and isinstance(celda_c2, str):
        texto = str(celda_c2).strip().upper()
        if "CLIENTES" in texto:
            partes = texto.split("CLIENTES")
            if len(partes) > 1:
                producto_detectado = partes[1].strip().split()[0]
        else:
            producto_detectado = texto.split()[-1]
    return producto_detectado

def procesar_archivo(ruta_archivo):
    """Procesa un único archivo Excel y devuelve un DataFrame limpio."""
    print(f"   procesando: {ruta_archivo.name}...")
    
    # 1. Metadatos (Fecha y Producto)
    fecha = np.nan
    producto = "DESCONOCIDO"
    try:
        wb = load_workbook(filename=ruta_archivo, data_only=True)
        # Fecha
        if HOJA_METADATOS in wb.sheetnames:
            val_fecha = wb[HOJA_METADATOS]["J3"].value
            if isinstance(val_fecha, datetime):
                fecha = val_fecha.date()
            else:
                try:
                    fecha = datetime.strptime(str(val_fecha).split()[0], "%d/%m/%Y").date()
                except:
                    pass
        # Producto
        producto = encontrar_producto_en_resultados(wb, HOJA_METADATOS)
    except Exception as e:
        print(f"   ⚠️ Advertencia leyendo metadatos en {ruta_archivo.name}: {e}")

    # 2. Datos
    try:
        df = pd.read_excel(ruta_archivo, sheet_name=HOJA_DATOS, usecols="A:D", header=3)
    except Exception as e:
        print(f"   ❌ Error leyendo hoja de datos en {ruta_archivo.name}: {e}")
        return None

    # Limpieza
    df.columns = df.columns.str.strip()
    df["Nombre proveedor"] = df["Nombre proveedor"].ffill().str.strip()
    df["Zona"] = df["Zona"].ffill().str.strip()
    df["Tipo Unidad"] = df["Tipo Unidad"].astype(str).str.strip()
    df["Kilos netos"] = pd.to_numeric(df["Kilos netos"], errors='coerce').fillna(0)
    
    # Filtrar
    df = df[df['Tipo Unidad'].isin(['Reclamo', 'Venta'])].copy()

    if df.empty:
        return None

    # 3. Pivot
    df_pivot = df.pivot_table(
        index=['Nombre proveedor', 'Zona'],
        columns='Tipo Unidad',
        values='Kilos netos',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    df_pivot.columns.name = None
    df_pivot = df_pivot.rename_axis(columns=None)
    
    # 4. Añadir columnas identificativas
    df_pivot.insert(0, "fecha", fecha)
    df_pivot.insert(1, "producto", producto)
    # Opcional: Añadir el nombre del archivo original por si quieres rastrear errores
    df_pivot["archivo_origen"] = ruta_archivo.name 

    return df_pivot

def main():
    print(f"--- Iniciando proceso MULTI-ARCHIVO en GitHub Actions ---")
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Buscamos TODOS los archivos .xlsx en la carpeta input
    lista_archivos = list(INPUT_DIR.glob("*.xlsx"))
    
    if not lista_archivos:
        print(f"❌ ERROR: No hay archivos .xlsx en la carpeta 'input'.")
        exit(1)

    print(f"📂 Se encontraron {len(lista_archivos)} archivos para procesar.")
    
    dfs_acumulados = []

    # Bucle para procesar cada archivo
    for archivo in lista_archivos:
        df_procesado = procesar_archivo(archivo)
        if df_procesado is not None:
            dfs_acumulados.append(df_procesado)
    
    # Unificación
    if dfs_acumulados:
        print("🔗 Unificando resultados...")
        df_final = pd.concat(dfs_acumulados, ignore_index=True)
        
        try:
            df_final.to_excel(RUTA_SALIDA, sheet_name='Datos_Unificados', index=False)
            print(f"✅ ÉXITO TOTAL: Archivo unificado guardado en: {RUTA_SALIDA}")
        except Exception as e:
            print(f"❌ Error al guardar el archivo unificado: {e}")
            exit(1)
    else:
        print("⚠️ No se generaron datos válidos de ninguno de los archivos.")

if __name__ == "__main__":
    main()
